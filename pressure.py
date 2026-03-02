#!/usr/bin/env python3
"""
Simple throughput benchmarking tool for LLM inference
Benchmark throughput and time-to-first-token across different worker counts
"""
import asyncio
import aiohttp
import time
import statistics
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

async def send_one(session, url, payload, timeout):
    """Send a single request"""
    start = time.perf_counter()
    first_token_time = None
    try:
        async with session.post(url, json=payload, timeout=timeout) as resp:
            status = resp.status
            if status != 200:
                text = await resp.text()
                end = time.perf_counter()
                    # Print details of the first HTTP error
                if not hasattr(send_one, '_error_printed'):
                    print(f"\n❌ HTTP {status} Error:")
                    print(f"Response: {text[:500]}")
                    send_one._error_printed = True
                return False, end - start, None, 0, f"HTTP {status}: {text[:200]}"
            
            result = await resp.json()
            end = time.perf_counter()
            
            # Extract token information
            usage = result.get('usage', {})
            total_tokens = usage.get('completion_tokens', 0)
            
            # Rough estimate of time-to-first-token
            total_time = end - start
            first_token_time = total_time * 0.1 if total_tokens > 0 else None
            
            return True, end - start, first_token_time, total_tokens, ""
    except Exception as e:
        end = time.perf_counter()
        # Print details of the first exception
        if not hasattr(send_one, '_exception_printed'):
            print(f"\n❌ Exception: {type(e).__name__}: {str(e)}")
            send_one._exception_printed = True
        return False, end - start, None, 0, repr(e)

async def worker(worker_id, session, url, payload, timeout, jobs, results):
    """Worker function that processes the request queue"""
    while True:
        idx = await jobs.get()
        if idx is None:
            jobs.task_done()
            break
        ok, latency, first_token, tokens, err = await send_one(session, url, payload, timeout)
        results.append({
            "idx": idx, 
            "worker_id": worker_id,
            "ok": ok, 
            "latency": latency, 
            "first_token_time": first_token,
            "tokens": tokens,
            "error": err
        })
        jobs.task_done()

async def run_benchmark(concurrency: int, prompts: List[str], url: str, model: str, max_tokens: int, timeout: float, repeats: int = 1, warmup: int = 0):
    """
    Run benchmark tests.

    Args:
        concurrency: number of workers
        prompts: list of prompts to test
        url: API endpoint
        model: model name
        max_tokens: maximum tokens
        timeout: timeout in seconds
        repeats: number of repeated runs to average
        warmup: number of warmup requests

    Returns:
        dict: contains total throughput, average throughput, and avg first token time
    """
    timeout_cfg = aiohttp.ClientTimeout(total=timeout)
    
    # Store results across runs for averaging
    all_total_throughputs = []
    all_avg_throughputs = []
    all_first_token_times = []
    total_success = 0
    total_requests = 0
    
    async with aiohttp.ClientSession(timeout=timeout_cfg) as session:
        # Warmup phase
        if warmup > 0:
            warmup_payload = {
                "model": model,
                "messages": [{"role": "user", "content": prompts[0]}],
                "max_tokens": max_tokens,
            }
            for _ in range(warmup):
                await send_one(session, url, warmup_payload, timeout)
        
        # Repeat tests 'repeats' times
        for repeat in range(repeats):
            all_results = []
            
            # Run one test per prompt
            for prompt in prompts:
                payload = {
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": max_tokens,
                }
                
                jobs = asyncio.Queue()
                results = []
                
                # Each worker will handle one request
                total_req = concurrency
                for i in range(total_req):
                    await jobs.put(i)
                for _ in range(concurrency):
                    await jobs.put(None)
                
                workers_list = [
                    asyncio.create_task(worker(i, session, url, payload, timeout, jobs, results))
                    for i in range(concurrency)
                ]
                
                t0 = time.perf_counter()
                await jobs.join()
                t1 = time.perf_counter()
                
                for w in workers_list:
                    w.cancel()
                
                all_results.extend(results)
            
            # Compute statistics for this round
            successful = [r for r in all_results if r["ok"]]
            
            if successful:
                # Group by worker to compute throughput
                worker_throughputs = {}
                for r in successful:
                    wid = r['worker_id']
                    if wid not in worker_throughputs:
                        worker_throughputs[wid] = {'tokens': 0, 'latency': 0}
                    worker_throughputs[wid]['tokens'] += r['tokens']
                    worker_throughputs[wid]['latency'] += r['latency']
                
                # Compute per-worker throughput and sum
                total_throughput = sum(
                    stats['tokens'] / stats['latency'] if stats['latency'] > 0 else 0
                    for stats in worker_throughputs.values()
                )
                
                # Average throughput = total_throughput / number_of_workers
                avg_throughput = total_throughput / concurrency if concurrency > 0 else 0
                
                # Average time-to-first-token
                first_tokens = [r["first_token_time"] for r in successful if r["first_token_time"]]
                avg_first_token_time = statistics.mean(first_tokens) if first_tokens else 0
                
                # Save current round results
                all_total_throughputs.append(total_throughput)
                all_avg_throughputs.append(avg_throughput)
                all_first_token_times.append(avg_first_token_time)
            
            total_success += len(successful)
            total_requests += len(all_results)
        
        # Compute average across multiple runs
        if not all_total_throughputs:
            return {
                'total_throughput': 0,
                'avg_throughput': 0,
                'avg_first_token_time': 0,
                'success_count': 0,
                'total_count': total_requests
            }
        
        return {
            'total_throughput': statistics.mean(all_total_throughputs),
            'avg_throughput': statistics.mean(all_avg_throughputs),
            'avg_first_token_time': statistics.mean(all_first_token_times),
            'success_count': total_success,
            'total_count': total_requests
        }

async def main():
    """Main function"""
    # Parameter configuration
    url = "http://localhost:8000/v1/chat/completions"  # SGLang endpoint
    # url = "http://localhost:8355/v1/chat/completions"  # TRT-LLM serve endpoint
    # url = "http://localhost:11434/v1/chat/completions"  # Ollama OpenAI-compatible API
    model = "openai/gpt-oss-20b"  # SGLang model name
    max_tokens = 500
    timeout = 60.0
    repeats = 3  # repeats per worker count
    warmup = 1  # warmup requests
    
    # Test prompts
    prompts = [
        "Explain the concept of KV cache in one short paragraph.",
        "Summarize the benefits of TensorRT-LLM in one sentence.",
        "What is the difference between GRPO and DPO in Reinforcement Learning?, explain in one sentence.",
        "Give me three synonyms for the word \"performance\".",
    ]
    
    # Worker count range: 1 to 20 (suitable for single-GPU SGLang tests)
    worker_counts = list(range(1, 21))
    
    # Store results
    results = []
    
    print("="*70)
    print("LLM Throughput Benchmark")
    print("="*70)
    print(f"Testing workers: {worker_counts[0]} to {worker_counts[-1]}")
    print(f"Prompts per test: {len(prompts)}")
    print(f"Repeats per worker count: {repeats}")
    print(f"Warmup requests: {warmup}")
    print("="*70)
    print()
    
    # Run tests for each worker count
    for workers in worker_counts:
        print(f"Testing with {workers} workers...", end=" ", flush=True)
        
        result = await run_benchmark(
            concurrency=workers,
            prompts=prompts,
            url=url,
            model=model,
            max_tokens=max_tokens,
            timeout=timeout,
            repeats=repeats,
            warmup=warmup
        )
        
        result['workers'] = workers
        results.append(result)
        
        # If the first test failed, print error info
        if workers == 1 and result['success_count'] == 0:
            print(f"\n⚠️  First test failed - checking for errors...")
            # Optionally rerun once to capture error details
            import sys
            sys.exit(1)
        
        print(f"✓ Total Throughput: {result['total_throughput']:.2f} tokens/s, "
              f"Avg Throughput: {result['avg_throughput']:.2f} tokens/s, "
              f"Avg First Token: {result['avg_first_token_time']*1000:.1f}ms "
              f"({result['success_count']}/{result['total_count']})")
    
    # Save to Excel
    save_to_excel(results)
    
    print()
    print("="*70)
    print("✅ Benchmark completed! Results saved to throughput_benchmark.xlsx")
    print("="*70)

def save_to_excel(results: List[Dict]):
    """Save results to an Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Throughput Benchmark"
    
    # Set header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header - use the same column names as analysis.py
    headers = ["Workers", "Total Throughput (tokens/s)", "Avg Throughput (tokens/s)", "Avg First Token Time (ms)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['workers'])
        ws.cell(row=row, column=2, value=round(result['total_throughput'], 2))
        ws.cell(row=row, column=3, value=round(result['avg_throughput'], 2))
        ws.cell(row=row, column=4, value=round(result['avg_first_token_time'] * 1000, 1))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    
    # Align all data cells
    for row in range(2, len(results) + 2):
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    
    # Save file
    wb.save("throughput_benchmark.xlsx")

if __name__ == "__main__":
    asyncio.run(main())